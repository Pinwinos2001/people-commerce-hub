import pandas as pd
from datetime import datetime, date
from supabase import create_client

SUPABASE_URL = "https://ditdfbipimipdmwptpcs.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRpdGRmYmlwaW1pcGRtd3B0cGNzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzE4OTU0MjMsImV4cCI6MjA4NzQ3MTQyM30.AeIAuv2t9eRo3aL0IyPwTY--uIXNiondrgmeJt8jPb8"

def get_client():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

def init_data():
    pass

# ── PENDIENTES ────────────────────────────────────────────────────────────────

def get_pendientes() -> pd.DataFrame:
    try:
        res = get_client().table("pendientes").select("*").order("id").execute()
        df = pd.DataFrame(res.data)
        if not df.empty and "deadline" in df.columns:
            df["deadline"] = pd.to_datetime(df["deadline"], errors="coerce")
        return df if not df.empty else _empty("pendientes")
    except:
        return _empty("pendientes")

def add_pendiente(tema, accion, responsable, deadline, status, prioridad, comentarios, creado_por):
    get_client().table("pendientes").insert({
        "tema": tema, "accion": accion, "responsable": responsable,
        "deadline": str(deadline) if deadline else None,
        "status": status, "prioridad": prioridad,
        "comentarios": comentarios, "creado_por": creado_por,
    }).execute()

def update_pendiente(row_id, campo, valor):
    get_client().table("pendientes").update({campo: valor}).eq("id", row_id).execute()

def delete_pendiente(row_id):
    get_client().table("pendientes").delete().eq("id", row_id).execute()

# ── PROCESOS ──────────────────────────────────────────────────────────────────

def get_procesos() -> pd.DataFrame:
    try:
        res = get_client().table("procesos").select("*").order("id").execute()
        df = pd.DataFrame(res.data)
        if not df.empty:
            for col in ["inicio_reclutamiento", "fecha_ingreso_plan", "fecha_ingreso_real"]:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
        return df if not df.empty else _empty("procesos")
    except:
        return _empty("procesos")

def add_proceso(data: dict) -> int:
    clean = {k: (str(v) if isinstance(v, date) else v) for k, v in data.items()}
    res = get_client().table("procesos").insert(clean).execute()
    return res.data[0]["id"] if res.data else None

def update_proceso(row_id, updates: dict):
    get_client().table("procesos").update(updates).eq("id", row_id).execute()

def cerrar_proceso(row_id, seleccionado, fecha_ingreso_real, tipo_cierre, reemplaza=None):
    res = get_client().table("procesos").select("*").eq("id", row_id).execute()
    if not res.data:
        return
    row = res.data[0]
    update_proceso(row_id, {
        "seleccionado": seleccionado,
        "fecha_ingreso_real": fecha_ingreso_real,
        "tipo_cierre": tipo_cierre,
        "status": "Cerrado",
        "reemplaza": reemplaza or row.get("reemplaza", ""),
    })
    add_headcount({
        "nombre": seleccionado, "cargo": row["posicion"], "area": row["area"],
        "fecha_ingreso": fecha_ingreso_real, "tipo_movimiento": "Ingreso",
        "activo": True, "comentarios": f"Proceso R&S #{row_id}",
    })
    if reemplaza and str(reemplaza).strip() and reemplaza != "NUEVO HC":
        add_cese({
            "nombre": reemplaza, "cargo": row["posicion"], "area": row["area"],
            "fecha_cese": fecha_ingreso_real, "motivo": "Baja (proceso de reemplazo)",
            "reemplazado_por": seleccionado, "proceso_id": row_id,
        })
        get_client().table("headcount").update({"activo": False, "fecha_cese": fecha_ingreso_real})\
            .ilike("nombre", reemplaza).execute()

# ── HEADCOUNT ─────────────────────────────────────────────────────────────────

def get_headcount() -> pd.DataFrame:
    try:
        res = get_client().table("headcount").select("*").order("id").execute()
        df = pd.DataFrame(res.data)
        if not df.empty:
            for col in ["fecha_ingreso", "fecha_cese"]:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
        return df if not df.empty else _empty("headcount")
    except:
        return _empty("headcount")

def add_headcount(data: dict):
    clean = {k: (str(v) if isinstance(v, date) else v) for k, v in data.items()}
    get_client().table("headcount").insert(clean).execute()

def get_ceses() -> pd.DataFrame:
    try:
        res = get_client().table("ceses").select("*").order("id").execute()
        df = pd.DataFrame(res.data)
        if not df.empty and "fecha_cese" in df.columns:
            df["fecha_cese"] = pd.to_datetime(df["fecha_cese"], errors="coerce")
        return df if not df.empty else _empty("ceses")
    except:
        return _empty("ceses")

def add_cese(data: dict):
    clean = {k: (str(v) if isinstance(v, date) else v) for k, v in data.items()}
    get_client().table("ceses").insert(clean).execute()

# ── HELPERS ───────────────────────────────────────────────────────────────────

def _empty(tabla: str) -> pd.DataFrame:
    cols = {
        "pendientes": ["id","tema","accion","responsable","deadline","status","prioridad","comentarios","creado_por","fecha_creacion"],
        "procesos":   ["id","año","posicion","area","responsable","jefe_directo","gerente","reemplaza","tipo_cobertura","vacantes","etapa","inicio_reclutamiento","seleccionado","fecha_ingreso_plan","fecha_ingreso_real","tipo_cierre","status","comentarios","fecha_creacion"],
        "headcount":  ["id","nombre","cargo","area","fecha_ingreso","fecha_cese","tipo_movimiento","motivo_cese","reemplazado_por","activo","comentarios","fecha_registro"],
        "ceses":      ["id","nombre","cargo","area","fecha_cese","motivo","reemplazado_por","proceso_id","fecha_registro"],
    }
    return pd.DataFrame(columns=cols.get(tabla, []))

# ── IMPORT EXCEL ──────────────────────────────────────────────────────────────

def import_headcount_excel(uploaded_file) -> int:
    df = pd.read_excel(uploaded_file)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    col_map = {
        "nombre":       ["nombre","name","apellidos_y_nombres","colaborador"],
        "cargo":        ["cargo","posicion","position","puesto"],
        "fecha_ingreso":["fecha_ingreso","ingreso","fecha_de_ingreso","start_date"],
        "fecha_cese":   ["fecha_cese","cese","fecha_de_cese","end_date","fecha_salida"],
    }
    rename = {}
    for std, options in col_map.items():
        for opt in options:
            if opt in df.columns:
                rename[opt] = std
                break
    df = df.rename(columns=rename)
    count = 0
    for _, row in df.iterrows():
        nombre = str(row.get("nombre","")).strip()
        if not nombre or nombre.lower() == "nan":
            continue
        fecha_cese = row.get("fecha_cese")
        activo = pd.isna(fecha_cese) or str(fecha_cese).strip() == ""
        add_headcount({
            "nombre": nombre,
            "cargo": str(row.get("cargo","Desarrollador de Ventas")),
            "area": "Ventas",
            "fecha_ingreso": str(row.get("fecha_ingreso","")) or None,
            "fecha_cese": None if activo else str(fecha_cese),
            "tipo_movimiento": "Ingreso",
            "activo": activo,
            "comentarios": "Importado desde Excel",
        })
        count += 1
    return count

# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────

def export_to_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    sheets_data = {
        "Pendientes":   get_pendientes(),
        "Procesos R&S": get_procesos(),
        "Headcount":    get_headcount(),
        "Ceses":        get_ceses(),
    }
    first = True
    for name, df in sheets_data.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        hf = PatternFill("solid", start_color="008C45", end_color="008C45")
        hfont = Font(bold=True, color="FFFFFF")
        for ci, cn in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=cn.upper())
            cell.font = hfont
            cell.fill = hf
            cell.alignment = Alignment(horizontal="center")
        for ri, row in df.iterrows():
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri+2, column=ci, value=str(val) if pd.notna(val) else "")
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf